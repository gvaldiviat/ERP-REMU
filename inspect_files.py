import glob
import os
import openpyxl

def check():
    print("--- Searching for xlsx files in workspace ---")
    files = glob.glob("*.xlsx")
    for f in sorted(files):
        size_kb = os.path.getsize(f) / 1024
        print(f"File: {f} ({size_kb:.2f} KB)")
        
    # Inspect listado_empleados_activos.xlsx if present
    f_act = "listado_empleados_activos.xlsx"
    if os.path.exists(f_act):
        wb = openpyxl.load_workbook(f_act, read_only=True)
        print(f"\nSheets in {f_act}: {wb.sheetnames}")
        sheet = wb.active
        # Let's inspect headers on Row 2
        rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        if rows:
            print(f"Headers in {f_act} (row 2):")
            print(list(rows[0])[:15])
            
    # Inspect listado_empleados_inactivos.xlsx if present
    f_inact = "listado_empleados_inactivos.xlsx"
    if os.path.exists(f_inact):
        wb = openpyxl.load_workbook(f_inact, read_only=True)
        print(f"\nSheets in {f_inact}: {wb.sheetnames}")
        sheet = wb.active
        rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        if rows:
            print(f"Headers in {f_inact} (row 2):")
            print(list(rows[0])[:15])

if __name__ == "__main__":
    check()
