import openpyxl

def inspect_raw():
    wb = openpyxl.load_workbook("Detalle proceso 1 del 2026-01 al 2026-01.xlsx", data_only=True)
    sheet = wb["Detalle"]
    
    headers = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
    
    for row in sheet.iter_rows(min_row=6, values_only=True):
        if not row or len(row) < 7 or not row[5]: continue
        rut = str(row[5]).replace(".", "").replace(" ", "").upper()
        if "10742396" in rut:
            print("=== RAW ROW VALUES ===")
            for idx, val in enumerate(row):
                if val is not None and val != 0 and val != "":
                    print(f"  {idx}: {headers[idx]} -> {val}")

if __name__ == "__main__":
    inspect_raw()
