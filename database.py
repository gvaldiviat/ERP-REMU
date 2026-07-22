import sqlite3

# Monkey-patch sqlite3.connect to automatically use uri=True and nolock=1 to prevent WSL2 bind-mount disk I/O errors
if not hasattr(sqlite3, "_original_connect"):
    sqlite3._original_connect = sqlite3.connect
    def _custom_connect(database, *args, **kwargs):
        if isinstance(database, str) and (database.endswith('.db') or 'remuneraciones.db' in database) and not database.startswith('file:'):
            db_path = database.replace('\\', '/')
            import sys
            if sys.platform == 'win32':
                database = f"file:{db_path}?nolock=1"
            else:
                database = f"file:{db_path}?vfs=unix-none"
            kwargs['uri'] = True
        conn = sqlite3._original_connect(database, *args, **kwargs)
        try:
            conn.execute("PRAGMA busy_timeout = 10000;")
            conn.execute("PRAGMA journal_mode = MEMORY;")
            conn.execute("PRAGMA synchronous = OFF;")
        except:
            pass
        return conn
    sqlite3.connect = _custom_connect

import openpyxl
import shutil
import tempfile

if not hasattr(openpyxl, "_original_load_workbook"):
    openpyxl._original_load_workbook = openpyxl.load_workbook
    def _custom_load_workbook(filename, *args, **kwargs):
        if isinstance(filename, str) and os.path.exists(filename):
            dir_name = os.path.dirname(filename)
            temp_path = os.path.join(dir_name, "~temp_" + os.path.basename(filename))
            try:
                shutil.copyfile(filename, temp_path)
                wb = openpyxl._original_load_workbook(temp_path, *args, **kwargs)
                return wb
            finally:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass
        return openpyxl._original_load_workbook(filename, *args, **kwargs)
    openpyxl.load_workbook = _custom_load_workbook

    # Patch openpyxl CustomFilterValueDescriptor to ignore validation errors on invalid Excel filter values
    try:
        import openpyxl.worksheet.filters as filters
        from openpyxl.descriptors.base import Convertible
        def _custom_set(self, instance, value):
            if isinstance(value, str):
                m = self.pattern.match(value)
                if not m:
                    self.expected_type = str
                elif "*" in value:
                    self.expected_type = str
            Convertible.__set__(self, instance, value)
        filters.CustomFilterValueDescriptor.__set__ = _custom_set
    except Exception as patch_err:
        print(f"Warning: Failed to patch openpyxl filters: {patch_err}")

import json
import os
import glob

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "remuneraciones.db")
EMPLOYEE_EXCEL = os.path.join(BASE_DIR, "listado_empleados.xlsx")

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    print("Initializing database...")
    conn = get_connection()
    cursor = conn.cursor()

    # Truncate tables instead of dropping them to prevent schema lock conflicts in WSL2/Docker
    for table in ["empleados", "rex_comparisons", "planilla_entradas", "parametros", "liquidaciones", 
                  "cargos_mapping", "proyectos_mapping", "vacaciones_ajustes", "vacaciones_reporte", 
                  "finiquitos_guardados", "alertas_auditoria"]:
        try:
            cursor.execute(f"DELETE FROM {table}")
        except sqlite3.OperationalError:
            pass
    # Keep liquidaciones_snapshots to preserve process comparisons history

    # Create planilla_entradas table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS planilla_entradas (
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        dias_trabajados INTEGER,
        sueldo_base REAL,
        bono_descanso REAL,
        bono_feriado REAL,
        bono_incentivo REAL,
        bono_responsabilidad REAL,
        bono_gestion REAL,
        bono_permanencia REAL,
        gratificacion REAL,
        diferencia_gratificacion REAL,
        colacion REAL,
        movilizacion REAL,
        pasajes REAL,
        traslados REAL,
        bono_estudios REAL,
        bono_fallecimiento REAL,
        apvi REAL,
        anticipo REAL,
        ccaf_credito REAL,
        ccaf_prestamo REAL,
        retencion_judicial REAL,
        prestamos_empresa REAL,
        seguro_complementario REAL,
        falp REAL,
        ahorro_afp REAL,
        licencia_dias INTEGER,
        justificaciones_json TEXT,
        observaciones TEXT,
        hash_origen TEXT,
        timestamp_carga TEXT,
        PRIMARY KEY (rut, contrato, periodo)
    )
    """)

    # Create employees table with composite primary key (rut, contrato)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS empleados (
        rut TEXT,
        contrato INTEGER,
        nombre TEXT,
        sexo TEXT,
        fecha_nacimiento TEXT,
        estado_civil TEXT,
        comuna TEXT,
        correo TEXT,
        telefono TEXT,
        banco TEXT,
        cuenta_banco TEXT,
        forma_pago TEXT,
        afp TEXT,
        cotizacion_afp REAL,
        isapre TEXT,
        moneda_isapre TEXT,
        cotizacion_uf REAL,
        cotizacion_pesos REAL,
        tramo_asig_fam TEXT,
        tipo_contrato TEXT,
        fecha_inicio_contrato TEXT,
        fecha_termino_contrato TEXT,
        sueldo_base INTEGER,
        cargo TEXT,
        centro_costo TEXT,
        sede TEXT,
        horas_semanales REAL,
        afecto_seguro_cesantia INTEGER,
        numero_hijos INTEGER,
        agrupacion TEXT,
        area TEXT,
        raw_json TEXT,
        id_obra TEXT,
        fecha_finiquito TEXT,
        PRIMARY KEY (rut, contrato)
    )
    """)

    # Create rex_comparisons table with composite primary key (rut, contrato, periodo)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS rex_comparisons (
        rut TEXT,
        contrato INTEGER,
        nombre TEXT,
        dias_trabajados INTEGER,
        sueldo_base REAL,
        bono_descanso REAL,
        bono_feriado REAL,
        bono_incentivo REAL,
        bono_responsabilidad REAL,
        bono_gestion REAL,
        bono_permanencia REAL,
        gratificacion REAL,
        diferencia_gratificacion REAL,
        colacion REAL,
        movilizacion REAL,
        pasajes REAL,
        traslados REAL,
        bono_estudios REAL,
        bono_fallecimiento REAL,
        apvi REAL,
        cotizacion_afp REAL,
        cotizacion_salud REAL,
        seguro_cesantia_trab REAL,
        impuesto REAL,
        total_descuentos REAL,
        sueldo_liquido REAL,
        alcance_liquido REAL,
        mutual REAL,
        sis REAL,
        seguro_cesantia_emp REAL,
        costo_empresa REAL,
        total_imponible REAL,
        afecto_afp REAL,
        afecto_cesantia REAL,
        afecto_impuesto REAL,
        afp TEXT,
        isapre TEXT,
        tipo_contrato TEXT,
        periodo TEXT,
        cargo TEXT,
        centro_costo TEXT,
        sede TEXT,
        fecha_inicio TEXT,
        fecha_termino TEXT,
        licencia_dias REAL DEFAULT 0.0,
        ias_vacaciones REAL DEFAULT 0.0,
        ias_anos_servicio REAL DEFAULT 0.0,
        ias_aviso REAL DEFAULT 0.0,
        justificaciones_json TEXT,
        anticipo REAL DEFAULT 0.0,
        ccaf_credito REAL DEFAULT 0.0,
        ccaf_prestamo REAL DEFAULT 0.0,
        retencion_judicial REAL DEFAULT 0.0,
        prestamos_empresa REAL DEFAULT 0.0,
        seguro_complementario REAL DEFAULT 0.0,
        falp REAL DEFAULT 0.0,
        PRIMARY KEY (rut, contrato, periodo)
    )
    """)

    # Create parameters table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS parametros (
        clave TEXT PRIMARY KEY,
        valor REAL,
        descripcion TEXT
    )
    """)

    # Create liquidations table (calculated results)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS liquidaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rut TEXT,
        contrato INTEGER,
        periodo TEXT, -- Format YYYY-MM
        dias_trabajados INTEGER,
        sueldo_base INTEGER,
        horas_extras INTEGER,
        monto_horas_extras INTEGER,
        bono_descanso INTEGER,
        bono_feriado INTEGER,
        bono_incentivo INTEGER,
        bono_responsabilidad INTEGER,
        bono_gestion INTEGER,
        bono_permanencia INTEGER,
        gratificacion INTEGER,
        colacion INTEGER,
        movilizacion INTEGER,
        pasajes INTEGER,
        traslados INTEGER,
        bono_estudios INTEGER,
        bono_fallecimiento INTEGER,
        total_imponible INTEGER,
        total_no_imponible INTEGER,
        total_haberes INTEGER,
        descuento_afp INTEGER,
        descuento_salud_total INTEGER,
        descuento_salud_obligatoria INTEGER,
        descuento_afc INTEGER,
        base_tributable INTEGER,
        descuento_impuesto INTEGER,
        total_descuentos INTEGER,
        sueldo_liquido INTEGER,
        alcance_liquido INTEGER,
        aporte_sis INTEGER,
        aporte_mutual INTEGER,
        aporte_afc INTEGER,
        costo_empresa INTEGER,
        fecha_calculo TEXT,
        licencia_dias INTEGER DEFAULT 0,
        ias_vacaciones INTEGER DEFAULT 0,
        ias_anos_servicio INTEGER DEFAULT 0,
        ias_aviso INTEGER DEFAULT 0,
        justificaciones_json TEXT,
        asignacion_familiar INTEGER DEFAULT 0,
        descuento_apvi INTEGER DEFAULT 0,
        descuento_anticipo INTEGER DEFAULT 0,
        descuento_ccaf_credito INTEGER DEFAULT 0,
        descuento_ccaf_prestamo INTEGER DEFAULT 0,
        descuento_retencion_judicial INTEGER DEFAULT 0,
        descuento_prestamos_empresa INTEGER DEFAULT 0,
        descuento_seguro_complementario INTEGER DEFAULT 0,
        descuento_falp INTEGER DEFAULT 0,
        total_descuentos_legales INTEGER DEFAULT 0,
        total_descuentos_otros INTEGER DEFAULT 0,
        ias_anticipo_finiquito INTEGER DEFAULT 0,
        ias_seguro_cesantia_recuperado INTEGER DEFAULT 0,
        FOREIGN KEY (rut, contrato) REFERENCES empleados(rut, contrato)
    )
    """)

    # Create Cargo and Project lookup tables
    cursor.execute("CREATE TABLE IF NOT EXISTS cargos_mapping (cargo TEXT PRIMARY KEY, generico TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS proyectos_mapping (centro_costo TEXT PRIMARY KEY, generico TEXT)")

    # Create liquidations snapshots table (archive of previous calculations)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS liquidaciones_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        dias_trabajados INTEGER,
        horas_extras INTEGER,
        monto_horas_extras INTEGER,
        bono_descanso INTEGER,
        bono_feriado INTEGER,
        bono_incentivo INTEGER,
        bono_responsabilidad INTEGER,
        bono_gestion INTEGER,
        bono_permanencia INTEGER,
        gratificacion INTEGER,
        colacion INTEGER,
        movilizacion INTEGER,
        pasajes INTEGER,
        traslados INTEGER,
        bono_estudios INTEGER,
        bono_fallecimiento INTEGER,
        total_imponible INTEGER,
        total_no_imponible INTEGER,
        total_haberes INTEGER,
        descuento_afp INTEGER,
        descuento_salud_total INTEGER,
        descuento_salud_obligatoria INTEGER,
        descuento_afc INTEGER,
        base_tributable INTEGER,
        descuento_impuesto INTEGER,
        total_descuentos INTEGER,
        sueldo_liquido INTEGER,
        alcance_liquido INTEGER,
        aporte_sis INTEGER,
        aporte_mutual INTEGER,
        aporte_afc INTEGER,
        costo_empresa INTEGER,
        fecha_calculo TEXT,
        licencia_dias INTEGER DEFAULT 0,
        ias_vacaciones INTEGER DEFAULT 0,
        ias_anos_servicio INTEGER DEFAULT 0,
        ias_aviso INTEGER DEFAULT 0,
        justificaciones_json TEXT,
        asignacion_familiar INTEGER DEFAULT 0,
        descuento_apvi INTEGER DEFAULT 0,
        descuento_anticipo INTEGER DEFAULT 0,
        descuento_ccaf_credito INTEGER DEFAULT 0,
        descuento_ccaf_prestamo INTEGER DEFAULT 0,
        descuento_retencion_judicial INTEGER DEFAULT 0,
        descuento_prestamos_empresa INTEGER DEFAULT 0,
        descuento_seguro_complementario INTEGER DEFAULT 0,
        descuento_falp INTEGER DEFAULT 0,
        total_descuentos_legales INTEGER DEFAULT 0,
        total_descuentos_otros INTEGER DEFAULT 0,
        ias_anticipo_finiquito INTEGER DEFAULT 0,
        ias_seguro_cesantia_recuperado INTEGER DEFAULT 0,
        snapshot_timestamp TEXT DEFAULT (datetime('now', 'localtime'))
    )
    """)

    # Create reconciliaciones table for persisting manual approvals
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reconciliaciones (
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        aprobado INTEGER DEFAULT 0,
        nota TEXT,
        usuario TEXT DEFAULT 'Administrador',
        fecha_aprobacion TEXT DEFAULT (datetime('now', 'localtime')),
        PRIMARY KEY (rut, contrato, periodo)
    )
    """)

    # Create vacaciones_ajustes table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS vacaciones_ajustes (
        rut TEXT,
        contrato INTEGER,
        dias_reales REAL DEFAULT 0.0,
        dias_tomados REAL DEFAULT 0.0,
        fecha_actualizacion TEXT,
        nota TEXT,
        PRIMARY KEY (rut, contrato)
    )
    """)

    # Create vacaciones_reporte table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS vacaciones_reporte (
        rut TEXT,
        contrato INTEGER,
        dias REAL,
        desde TEXT,
        hasta TEXT,
        tipo TEXT,
        dia_type TEXT,
        periodo TEXT
    )
    """)

    # Create finiquitos_guardados table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS finiquitos_guardados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rut TEXT,
        contrato INTEGER,
        fecha_termino TEXT,
        causal TEXT,
        aviso_previo INTEGER DEFAULT 0,
        dias_vacaciones_pendientes REAL DEFAULT 0.0,
        ias_monto INTEGER DEFAULT 0,
        aviso_monto INTEGER DEFAULT 0,
        vacaciones_monto INTEGER DEFAULT 0,
        descuento_afc_monto INTEGER DEFAULT 0,
        total_finiquito INTEGER DEFAULT 0,
        fecha_calculo TEXT,
        nota TEXT,
        sueldo_base INTEGER DEFAULT 0,
        gratificacion INTEGER DEFAULT 0,
        movilizacion INTEGER DEFAULT 0,
        renta_1 INTEGER DEFAULT 0,
        renta_2 INTEGER DEFAULT 0,
        dias_periodo INTEGER DEFAULT 0,
        vac_devengadas REAL DEFAULT 0.0,
        vac_progresivas REAL DEFAULT 0.0,
        vac_inhabiles REAL DEFAULT 0.0,
        vac_tomadas REAL DEFAULT 0.0,
        valor_dia_vac REAL DEFAULT 0.0,
        indem_tiempo_servido_yn TEXT DEFAULT 'NO',
        tiempo_servido_meses REAL DEFAULT 0.0,
        tiempo_servido_monto INTEGER DEFAULT 0,
        years_servicio INTEGER DEFAULT 0,
        years_a_pagar INTEGER DEFAULT 0,
        valor_dia_ias REAL DEFAULT 0.0,
        compensatoria_monto INTEGER DEFAULT 0,
        prestamo_monto INTEGER DEFAULT 0,
        bono_1 INTEGER DEFAULT 0,
        bono_2 INTEGER DEFAULT 0,
        FOREIGN KEY (rut, contrato) REFERENCES empleados (rut, contrato)
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS alertas_auditoria (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        tipo_alerta TEXT,
        detalle TEXT,
        fecha_deteccion TEXT DEFAULT (datetime('now', 'localtime')),
        UNIQUE (rut, contrato, periodo, tipo_alerta, detalle)
    )
    """)

    # 1. Tabla kpi_desviaciones_mensuales
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS kpi_desviaciones_mensuales (
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        alcance_liquido_calc INTEGER,
        alcance_liquido_rex INTEGER,
        diff_alcance INTEGER,
        total_imponible_calc INTEGER,
        total_imponible_rex INTEGER,
        diff_imponible INTEGER,
        costo_empresa_calc INTEGER,
        costo_empresa_rex INTEGER,
        diff_costo INTEGER,
        reconciliado INTEGER DEFAULT 0,
        PRIMARY KEY (rut, contrato, periodo)
    )
    """)

    # 2. Columna virtual anio_fiscal
    try:
        cursor.execute("ALTER TABLE kpi_desviaciones_mensuales ADD COLUMN anio_fiscal INTEGER GENERATED ALWAYS AS (CAST(SUBSTR(periodo, 1, 4) AS INTEGER))")
    except sqlite3.OperationalError:
        pass

    # 3. Tabla logs_alertas
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS logs_alertas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rut TEXT,
        contrato INTEGER,
        tipo_alerta TEXT,
        descripcion TEXT,
        periodos_afectados TEXT,
        deriva_acumulada INTEGER,
        fecha_creacion TEXT DEFAULT (datetime('now', 'localtime')),
        leida INTEGER DEFAULT 0,
        nota_resolucion TEXT,
        FOREIGN KEY(rut, contrato) REFERENCES empleados(rut, contrato)
    )
    """)

    # 4. Índices
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_kpi_desv_unique ON kpi_desviaciones_mensuales(rut, periodo, contrato)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_kpi_desv_anio_activo ON kpi_desviaciones_mensuales(anio_fiscal) WHERE reconciliado = 0")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_logs_alertas_no_leidas ON logs_alertas(leida, fecha_creacion DESC)")

    # 5. Triggers
    cursor.execute("""
    CREATE TRIGGER IF NOT EXISTS trg_sync_desviaciones_insert
    AFTER INSERT ON liquidaciones
    BEGIN
        INSERT OR REPLACE INTO kpi_desviaciones_mensuales
        SELECT 
            new.rut,
            new.contrato,
            new.periodo,
            new.alcance_liquido,
            COALESCE(c.alcance_liquido, 0),
            (new.alcance_liquido - COALESCE(c.alcance_liquido, 0)),
            new.total_imponible,
            COALESCE(c.total_imponible, 0),
            (new.total_imponible - COALESCE(c.total_imponible, 0)),
            new.costo_empresa,
            COALESCE(c.costo_empresa, 0),
            (new.costo_empresa - COALESCE(c.costo_empresa, 0)),
            COALESCE((SELECT aprobado FROM reconciliaciones WHERE rut = new.rut AND contrato = new.contrato AND periodo = new.periodo), 0)
        FROM rex_comparisons c
        WHERE c.rut = new.rut AND c.contrato = new.contrato AND c.periodo = new.periodo;
    END;
    """)

    cursor.execute("""
    CREATE TRIGGER IF NOT EXISTS trg_sync_reconciliacion_upsert
    AFTER INSERT ON reconciliaciones
    BEGIN
        UPDATE kpi_desviaciones_mensuales
        SET reconciliado = new.aprobado
        WHERE rut = new.rut 
          AND contrato = new.contrato 
          AND periodo = new.periodo;
    END;
    """)

    cursor.execute("""
    CREATE TRIGGER IF NOT EXISTS trg_sync_reconciliacion_update
    AFTER UPDATE ON reconciliaciones
    BEGIN
        UPDATE kpi_desviaciones_mensuales
        SET reconciliado = new.aprobado
        WHERE rut = new.rut 
          AND contrato = new.contrato 
          AND periodo = new.periodo;
    END;
    """)

    # 6. Vista v_auditoria_deriva_acumulada
    cursor.execute("""
    CREATE VIEW IF NOT EXISTS v_auditoria_deriva_acumulada AS
    SELECT 
        rut,
        contrato,
        periodo,
        diff_alcance,
        SUM(diff_alcance) OVER (
            PARTITION BY rut, contrato 
            ORDER BY periodo ASC 
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS deriva_acumulada_alcance,
        diff_costo,
        SUM(diff_costo) OVER (
            PARTITION BY rut, contrato 
            ORDER BY periodo ASC 
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS deriva_acumulada_costo
    FROM kpi_desviaciones_mensuales;
    """)

    # 7. Vista v_costo_no_calidad
    cursor.execute("""
    CREATE VIEW IF NOT EXISTS v_costo_no_calidad AS
    SELECT 
        periodo,
        COUNT(DISTINCT rut) AS total_empleados_afectados,
        SUM(ABS(diff_alcance)) AS fuga_neta_liquido_detectada,
        SUM(ABS(diff_costo)) AS desviacion_presupuestaria_total,
        SUM(CASE WHEN reconciliado = 1 THEN ABS(diff_costo) ELSE 0 END) AS costo_mitigado,
        SUM(CASE WHEN reconciliado = 0 THEN ABS(diff_costo) ELSE 0 END) AS riesgo_financiero_activo
    FROM kpi_desviaciones_mensuales
    GROUP BY periodo;
    """)

    # 8. Vista v_riesgo_provision_finiquitos
    cursor.execute("""
    CREATE VIEW IF NOT EXISTS v_riesgo_provision_finiquitos AS
    SELECT 
        e.rut,
        e.contrato,
        e.nombre,
        e.id_obra,
        e.sueldo_base,
        e.fecha_inicio_contrato,
        CAST((strftime('%s', '2026-06-30') - strftime('%s', e.fecha_inicio_contrato)) / (3600 * 24 * 30.43) AS INTEGER) AS meses_antiguedad,
        ROUND(CAST((strftime('%s', '2026-06-30') - strftime('%s', e.fecha_inicio_contrato)) / (3600 * 24 * 30.43) AS REAL) * 1.25, 2) AS vacaciones_proporcionales_estimadas,
        ROUND(e.sueldo_base / 30.0, 2) AS valor_dia_estimado,
        ROUND(
            (MIN(11, CAST((strftime('%s', '2026-06-30') - strftime('%s', e.fecha_inicio_contrato)) / (3600 * 24 * 365) AS INTEGER)) * e.sueldo_base) + 
            (CAST((strftime('%s', '2026-06-30') - strftime('%s', e.fecha_inicio_contrato)) / (3600 * 24 * 30.43) AS REAL) * 1.25 * (e.sueldo_base / 30.0)),
            0
        ) AS pasivo_laboral_estimado
    FROM empleados e
    WHERE e.fecha_finiquito IS NULL OR e.fecha_finiquito = '';
    """)

    conn.commit()
    conn.close()
    print("Database tables created.")

def seed_parameters():
    print("Seeding monthly parameters for May 2026...")
    default_params = [
        ("uf", 40610.69, "Valor UF a fines de Mayo 2026"),
        ("utm", 70588.00, "Valor UTM Mayo 2026"),
        ("imm", 539000.00, "Ingreso Mínimo Mensual para trabajadores activos (18-65 años)"),
        ("imm_otros", 402082.00, "Ingreso Mínimo Mensual para menores de 18 y mayores de 65 años"),
        ("tope_imponible_afp_uf", 90.0, "Tope Imponible AFP y Salud (en UF)"),
        ("tope_imponible_afc_uf", 135.2, "Tope Imponible Seguro Cesantía (en UF)"),
        ("sis_tasa", 1.62, "Seguro de Invalidez y Sobrevivencia % (SIS)"),
        ("mutual_tasa", 0.93, "Tasa Mutualidad Básica % (e.g. ACHS/Mutual, configurable)"),
        ("sanna_tasa", 0.03, "Tasa Ley SANNA %"),
    ]

    conn = get_connection()
    cursor = conn.cursor()
    for clave, valor, desc in default_params:
        cursor.execute("INSERT OR IGNORE INTO parametros (clave, valor, descripcion) VALUES (?, ?, ?)", (clave, valor, desc))
    conn.commit()
    conn.close()
    print("Monthly parameters seeded.")

def clean_rut(rut_str):
    if not rut_str:
        return ""
    return str(rut_str).replace(".", "").replace(" ", "").upper()

def format_date(dt):
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d")
    s = str(dt).strip()
    if not s:
        return ""
    import re
    # 1. Match DD/MM/YYYY or DD-MM-YYYY (with optional time or space)
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})', s)
    if m:
        d, m_val, y = m.groups()
        return f"{y}-{int(m_val):02d}-{int(d):02d}"
    # 2. Match YYYY/MM/DD or YYYY-MM-DD
    m2 = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})', s)
    if m2:
        y, m_val, d = m2.groups()
        return f"{y}-{int(m_val):02d}-{int(d):02d}"
    return s[:10]

def load_employees_from_excel(custom_paths=None):
    active_path = os.path.join(BASE_DIR, "listado_empleados_activos.xlsx")
    inactive_path = os.path.join(BASE_DIR, "listado_empleados_inactivos.xlsx")
    fallback_path = os.path.join(BASE_DIR, "listado_empleados.xlsx")
    
    paths_to_load = []
    if custom_paths:
        paths_to_load = custom_paths
    else:
        if os.path.exists(active_path):
            paths_to_load.append((active_path, "Activos"))
        if os.path.exists(inactive_path):
            paths_to_load.append((inactive_path, "Inactivos"))
            
        if not paths_to_load:
            if os.path.exists(fallback_path):
                paths_to_load.append((fallback_path, "Estándar"))
            else:
                print("Error: No employee Excel files found!")
                return
            
    conn = get_connection()
    cursor = conn.cursor()
    total_count = 0
    
    # Ensure new columns exist in case table wasn't dropped
    try:
        cursor.execute("ALTER TABLE empleados ADD COLUMN agrupacion TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE empleados ADD COLUMN area TEXT")
    except sqlite3.OperationalError:
        pass
    
    for path, label in paths_to_load:
        print(f"Loading employees from Excel {os.path.basename(path)} ({label})...")
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active

        # Row 2 has headers
        headers = [cell for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        
        def clean(s):
            if not s: return ""
            s = str(s).lower().strip()
            s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ñ", "n")
            return s
            
        def get_col_idx(headers_row, candidates):
            cleaned_candidates = [clean(c) for c in candidates]
            for idx, h in enumerate(headers_row):
                if h and clean(h) in cleaned_candidates:
                    return idx
            # Búsqueda parcial por si viene con sufijos extra
            for idx, h in enumerate(headers_row):
                if h:
                    ch = clean(h)
                    for cc in cleaned_candidates:
                        if cc and ch and (cc in ch or ch in cc):
                            return idx
            return None

        idx_rut = get_col_idx(headers, ["rut", "r.u.t"])
        idx_nombre = get_col_idx(headers, ["nombre", "trabajador"])
        idx_contrato = get_col_idx(headers, ["contrato", "n° contrato", "nro contrato"])
        idx_sexo = get_col_idx(headers, ["sexo", "genero"])
        idx_fecha_nac = get_col_idx(headers, ["fecha nacimiento", "nacimiento"])
        idx_est_civil = get_col_idx(headers, ["estado civil", "civil"])
        idx_comuna = get_col_idx(headers, ["comuna", "ciudad"])
        idx_correo = get_col_idx(headers, ["correo", "email", "e-mail"])
        idx_telefono = get_col_idx(headers, ["telefono", "celular"])
        idx_banco = get_col_idx(headers, ["banco", "institucion bancaria"])
        idx_cuenta_banco = get_col_idx(headers, ["cuenta banco", "n° cuenta", "numero cuenta"])
        idx_forma_pago = get_col_idx(headers, ["forma pago", "tipo pago"])
        idx_afp = get_col_idx(headers, ["afp", "institucion previsional"])
        idx_cotiz_afp = get_col_idx(headers, ["cotizacion afp", "tasa afp", "% afp"])
        idx_isapre = get_col_idx(headers, ["isapre", "institucion salud", "salud"])
        idx_moneda_isapre = get_col_idx(headers, ["moneda isapre", "moneda plan", "moneda salud"])
        idx_cotiz_uf = get_col_idx(headers, ["cotizacion uf", "plan uf"])
        idx_cotiz_pesos = get_col_idx(headers, ["cotizacion pesos", "plan pesos"])
        idx_tramo_asig = get_col_idx(headers, ["tramo asig", "tramo familiar"])
        idx_tipo_contrato = get_col_idx(headers, ["tipo contrato"])
        idx_f_inicio = get_col_idx(headers, ["fecha inicio contrato", "fecha ingreso contrato", "fecha inicio", "fecha ingreso"])
        idx_f_termino = get_col_idx(headers, ["fecha termino contrato", "fecha retiro contrato", "fecha termino", "fecha retiro"])
        idx_sueldo_base = get_col_idx(headers, ["sueldo base"])
        idx_cargo = get_col_idx(headers, ["cargo", "rol", "puesto"])
        idx_cc = get_col_idx(headers, ["centro costo", "cc"])
        idx_sede = get_col_idx(headers, ["sede", "sucursal"])
        idx_hrs_semanales = get_col_idx(headers, ["horas semanales", "jornada"])
        idx_afecto_afc = get_col_idx(headers, ["afecto seguro cesantia", "afecto afc", "seguro cesantia"])
        idx_num_hijos = get_col_idx(headers, ["numero hijos", "cargas", "hijos"])
        idx_agrupacion = get_col_idx(headers, ["agrupacion", "agrupación"])
        idx_area = get_col_idx(headers, ["area", "área"])

        # Fallbacks (En caso de archivo sin cabeceras estándar, vuelve a los índices originales)
        if idx_rut is None: idx_rut = 0
        if idx_nombre is None: idx_nombre = 1
        if idx_contrato is None: idx_contrato = 56
        if idx_sueldo_base is None: idx_sueldo_base = 66

        count = 0
        # Data starts from row 3
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or idx_rut >= len(row) or not row[idx_rut]: # Skip if RUT is empty
                continue
                
            def val(idx, default):
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return row[idx]
                return default

            raw_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    val_cell = row[idx]
                    if hasattr(val_cell, "isoformat"):
                        val_cell = val_cell.isoformat()
                    raw_dict[header] = val_cell

            rut = clean_rut(val(idx_rut, ""))
            contrato = int(val(idx_contrato, 1))
            nombre = val(idx_nombre, "")
            sexo = val(idx_sexo, "")
            fecha_nac = format_date(val(idx_fecha_nac, ""))
            est_civil = val(idx_est_civil, "")
            comuna = val(idx_comuna, "")
            correo = val(idx_correo, "")
            telefono = val(idx_telefono, "")
            banco = val(idx_banco, "")
            cuenta_banco = str(val(idx_cuenta_banco, ""))
            forma_pago = val(idx_forma_pago, "")
            afp = str(val(idx_afp, "")).lower()
            cotiz_afp = val(idx_cotiz_afp, 0.0)
            isapre = str(val(idx_isapre, "")).lower()
            moneda_isapre = val(idx_moneda_isapre, "")
            cotiz_uf = val(idx_cotiz_uf, 0.0)
            cotiz_pesos = val(idx_cotiz_pesos, 0.0)
            tramo_asig = val(idx_tramo_asig, "D")
            tipo_contrato = val(idx_tipo_contrato, "")
            f_inicio = format_date(val(idx_f_inicio, ""))
            f_termino = format_date(val(idx_f_termino, ""))
            sueldo_base = int(val(idx_sueldo_base, 0))
            cargo = str(val(idx_cargo, "")).strip().upper()
            cc = val(idx_cc, "")
            sede = val(idx_sede, "")
            hrs_semanales = val(idx_hrs_semanales, 40.0)
            
            val_afc = val(idx_afecto_afc, 1)
            afecto_afc = 1 if (val_afc is True or str(val_afc).lower() in ("true", "1", "s", "si")) else 0
            num_hijos = int(val(idx_num_hijos, 0))
            agrupacion = str(val(idx_agrupacion, "")).strip()
            area = str(val(idx_area, "")).strip()

            raw_json = json.dumps(raw_dict, ensure_ascii=False)

            # Resolve id_obra
            id_obra = ""
            for k, v in raw_dict.items():
                k_clean = str(k).lower().strip().replace("_", " ").replace(".", "")
                if "id centro de costo" in k_clean or "id centro costo" in k_clean or "id cc" in k_clean or "id_obra" in k_clean or "id obra" in k_clean:
                    id_obra = str(v) if v is not None else ""
                    break
            
            fecha_finiquito = f_termino if f_termino else None

            cursor.execute("""
            INSERT OR REPLACE INTO empleados (
                rut, contrato, nombre, sexo, fecha_nacimiento, estado_civil, comuna, correo, telefono,
                banco, cuenta_banco, forma_pago, afp, cotizacion_afp, isapre, moneda_isapre,
                cotizacion_uf, cotizacion_pesos, tramo_asig_fam, tipo_contrato,
                fecha_inicio_contrato, fecha_termino_contrato, sueldo_base, cargo,
                centro_costo, sede, horas_semanales, afecto_seguro_cesantia, numero_hijos, agrupacion, area, raw_json,
                id_obra, fecha_finiquito
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rut, contrato, nombre, sexo, fecha_nac, est_civil, comuna, correo, telefono,
                banco, cuenta_banco, forma_pago, afp, cotiz_afp, isapre, moneda_isapre,
                cotiz_uf, cotiz_pesos, tramo_asig, tipo_contrato,
                f_inicio, f_termino, sueldo_base, cargo,
                cc, sede, hrs_semanales, afecto_afc, num_hijos, agrupacion, area, raw_json,
                id_obra, fecha_finiquito
            ))
            count += 1
            
        print(f"Loaded {count} employees from {os.path.basename(path)}.")
        total_count += count

    conn.commit()
    conn.close()
    print(f"Loaded total of {total_count} employees (contracts) into database from new sources.")

def load_cargos_mapping():
    print("Loading cargo mappings from CARGOS.xlsx...")
    cargos_path = os.path.join(BASE_DIR, "CARGOS.xlsx")
    if not os.path.exists(cargos_path):
        print("Warning: CARGOS.xlsx not found!")
        return

    wb = openpyxl.load_workbook(cargos_path, data_only=True)
    sheet = wb.active
    
    conn = get_connection()
    cursor = conn.cursor()
    
    count = 0
    # Row 1 is header: ('Cargo', 'Generico')
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        cargo = str(row[0]).strip().upper()
        generico = str(row[1]).strip()
        cursor.execute("INSERT OR REPLACE INTO cargos_mapping (cargo, generico) VALUES (?, ?)", (cargo, generico))
        count += 1
        
    conn.commit()
    conn.close()
    print(f"Loaded {count} cargo mappings.")

def load_proyectos_mapping():
    print("Loading project mappings from Proyectos.xlsx...")
    proyectos_path = os.path.join(BASE_DIR, "Proyectos.xlsx")
    if not os.path.exists(proyectos_path):
        print("Warning: Proyectos.xlsx not found!")
        return

    wb = openpyxl.load_workbook(proyectos_path, data_only=True)
    sheet = wb.active
    
    conn = get_connection()
    cursor = conn.cursor()
    
    count = 0
    # Row 1 is header: ('Centro Costo', 'Generico')
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        cc = str(row[0]).strip()
        generico = str(row[1]).strip()
        cursor.execute("INSERT OR REPLACE INTO proyectos_mapping (centro_costo, generico) VALUES (?, ?)", (cc, generico))
        count += 1
        
    conn.commit()
    conn.close()
    print(f"Loaded {count} project mappings.")

def load_all_justifications():
    print("Loading justifications from PLANILLA GENERAL spreadsheets...")
    all_extra_data = {} # maps period -> { (rut, contrato): {justs, observations, licencia_dias} }
    
    # 1. Load May (2026-05)
    may_extra = {}
    may_path = os.path.join(BASE_DIR, "PLANILLA GENERAL MAYO 2026.xlsx")
    if os.path.exists(may_path):
        try:
            wb = openpyxl.load_workbook(may_path, data_only=True)
            sheet = wb["Detalle"]
            headers = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
            
            just_mappings = {}
            for idx, h in enumerate(headers):
                if h == 'justificacion' and idx > 0:
                    concept_name = headers[idx-1]
                    just_mappings[idx] = concept_name
            
            for row in sheet.iter_rows(min_row=6, values_only=True):
                if not row or len(row) < 7 or not row[5]: continue
                rut = clean_rut(row[5])
                contrato = int(row[6]) if row[6] is not None else 1
                lic_dias = int(row[21]) if row[21] is not None else 0
                obs = row[99] if len(row) > 99 else ""
                
                justs = {}
                for just_idx, concept in just_mappings.items():
                    val = row[just_idx]
                    if val:
                        justs[concept] = str(val).strip()
                
                key = (rut, contrato)
                may_extra[key] = {
                    "justificaciones": justs,
                    "licencia_dias": lic_dias,
                    "observaciones": obs
                }
            print(f"Parsed {len(may_extra)} employee justifications for May 2026.")
        except Exception as e:
            print(f"Error parsing PLANILLA GENERAL MAYO 2026.xlsx: {e}")
    else:
        print("Warning: PLANILLA GENERAL MAYO 2026.xlsx not found!")
    
    all_extra_data['2026-05'] = may_extra
    
    # 2. Historical periods: Jan, Feb, Mar, Apr
    hist_periods = {
        '2026-01': (os.path.join(BASE_DIR, 'Datos de entrada', 'PLANILLA GENERAL ENERO 2026.xlsx'), 'Planilla General Enero 2026'),
        '2026-02': (os.path.join(BASE_DIR, 'Datos de entrada', 'PLANILLA GENERAL FEBRERO 2026.xlsx'), 'PLANILLA GENERAL FEBRERO 2026'),
        '2026-03': (os.path.join(BASE_DIR, 'Datos de entrada', 'PLANILLA GENERAL MARZO 2026.xlsx'), 'PLANILA GENERAL MARZO 2026'),
        '2026-04': (os.path.join(BASE_DIR, 'Datos de entrada', 'PLANILLA GENERAL ABRIL 2026.xlsx'), 'PLANILLA GENERAL ABRIL 2026'),
    }
    
    for period, (filename, sheetname) in hist_periods.items():
        if not os.path.exists(filename):
            print(f"Warning: {filename} not found, skipping justifications.")
            continue
        try:
            wb = openpyxl.load_workbook(filename, data_only=True)
            sheet = wb[sheetname]
            
            # Row 1 has headers
            headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            
            rut_idx = None
            for idx, h in enumerate(headers):
                if h and 'rut' in str(h).lower():
                    rut_idx = idx
                    break
            if rut_idx is None:
                rut_idx = 1 # fallback
                
            just_mappings = {}
            for idx, h in enumerate(headers):
                if h and ('justif' in str(h).lower() or 'observa' in str(h).lower()):
                    concept = str(h)
                    if "10-29" in concept or "1029" in concept:
                        concept = "Bono Compensatorio descanso"
                    elif "1034" in concept:
                        concept = "Bono Incentivo variable"
                    elif "1140" in concept:
                        concept = "Bono Responsabilidad"
                    elif "1092" in concept:
                        concept = "Diferencia de haber"
                    elif "colacion" in concept.lower():
                        concept = "Colacion"
                    elif "movilizaci" in concept.lower():
                        concept = "Movilizacion"
                    elif "viatico" in concept.lower():
                        concept = "Viatico"
                    elif "asignaci" in concept.lower():
                        concept = "Otras Asignaciones N/I"
                    elif "observaciones generales" in concept.lower():
                        concept = "Observaciones Generales"
                    just_mappings[idx] = concept
                    
            period_data = {}
            # Data starts from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= rut_idx or not row[rut_idx]:
                    continue
                
                rut = clean_rut(row[rut_idx])
                contrato = 1 # Default contract 1 for history
                
                justs = {}
                obs = ""
                for just_idx, concept in just_mappings.items():
                    val = row[just_idx]
                    if val:
                        if concept == "Observaciones Generales":
                            obs = str(val).strip()
                        else:
                            justs[concept] = str(val).strip()
                            
                key = (rut, contrato)
                period_data[key] = {
                    "justificaciones": justs,
                    "observaciones": obs,
                    "licencia_dias": 0
                }
            all_extra_data[period] = period_data
            print(f"Loaded {len(period_data)} justifications for period {period}")
        except Exception as e:
            print(f"Error loading justifications for {period}: {e}")
            
    return all_extra_data

def load_planilla_entradas(custom_files=None):
    print("Loading general monthly planilla inputs...")
    import hashlib
    import datetime
    
    if custom_files:
        files = custom_files
    else:
        # Default files to scan
        files = [
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILA GENERAL JUNIO 2026.xlsx"),
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILLA GENERAL MAYO 2026.xlsx"),
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILLA GENERAL ABRIL 2026.xlsx"),
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILLA GENERAL MARZO 2026.xlsx"),
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILLA GENERAL FEBRERO 2026.xlsx"),
            os.path.join(BASE_DIR, "Datos de entrada", "PLANILLA GENERAL ENERO 2026.xlsx")
        ]
        files = [f for f in files if os.path.exists(f)]

    if not files:
        print("No general monthly planilla files found.")
        return

    conn = get_connection()
    cursor = conn.cursor()
    
    all_extra_data = load_all_justifications()
    timestamp_now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_count = 0
    for filepath in files:
        print(f"Parsing Planilla Entrada: {os.path.basename(filepath)}")
        
        # Calculate SHA-256 hash
        sha256 = hashlib.sha256()
        with open(filepath, 'rb') as f:
            while True:
                data = f.read(65536)
                if not data:
                    break
                sha256.update(data)
        file_hash = sha256.hexdigest()

        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Search for sheet
        sheetname = None
        for s in ['Planilla General', 'Planilla General Enero 2026', 'PLANILLA GENERAL FEBRERO 2026', 'PLANILA GENERAL MARZO 2026', 'PLANILLA GENERAL ABRIL 2026', 'Detalle']:
            if s in wb.sheetnames:
                sheetname = s
                break
        sheet = wb[sheetname] if sheetname else wb.active

        # Dynamically find the header row by searching the first 15 rows for "rut" and "contrato"
        header_row_idx = 5
        for idx in range(1, 16):
            rows_iter = list(sheet.iter_rows(min_row=idx, max_row=idx, values_only=True))
            if rows_iter and rows_iter[0]:
                row_vals_clean = [str(c).lower().strip() for c in rows_iter[0] if c is not None]
                if any("rut" in c for c in row_vals_clean) and (any("contrato" in c for c in row_vals_clean) or any("cargo" in c for c in row_vals_clean) or any("apellido" in c for c in row_vals_clean) or any("nombre" in c for c in row_vals_clean)):
                    header_row_idx = idx
                    break
                    
        headers_row = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        
        def clean(s):
            if not s:
                return ""
            s = str(s).lower().strip()
            s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
            s = s.replace("ñ", "n").replace("ü", "u")
            s = s.replace("  ", " ")
            return s

        # Find the last 'Agrupación' column index to parse only user inputs (ignoring system summaries before it)
        idx_last_agrupacion = None
        for idx, h in enumerate(headers_row):
            if h and "agrup" in clean(h):
                idx_last_agrupacion = idx

        def get_col_idx(candidates, after_idx=None):
            cleaned_candidates = [clean(c) for c in candidates]
            for idx, h in enumerate(headers_row):
                if after_idx is not None and idx <= after_idx:
                    continue
                if h and clean(h) in cleaned_candidates:
                    return idx
            for idx, h in enumerate(headers_row):
                if after_idx is not None and idx <= after_idx:
                    continue
                if h:
                    ch = clean(h)
                    for cc in cleaned_candidates:
                        if cc == "contrato" and any(term in ch for term in ["inicio", "termino", "tipo", "fin", "ingreso", "retiro"]):
                            continue
                        if cc and ch:
                            if len(ch) <= 3 and ch != cc:
                                continue
                            if cc in ch or ch in cc:
                                return idx
            return None

        idx_rut = get_col_idx(["rut"])
        idx_contrato = get_col_idx(["contrato"])
        idx_dias = get_col_idx(["dias trabajados", "das trabajados", "trabajados", "tdt", "t.d.t."])
        idx_sueldo_base = get_col_idx(["sueldo base contrato", "sueldo base pactado", "sueldo base"])
        
        idx_bono_desc = get_col_idx(["bono descanso", "descanso"], after_idx=idx_last_agrupacion)
        idx_bono_feri = get_col_idx(["bono feriado", "feriado"], after_idx=idx_last_agrupacion)
        idx_bono_inc = get_col_idx(["bono incentivo variable", "bono incentivo", "incentivo"], after_idx=idx_last_agrupacion)
        idx_bono_resp = get_col_idx(["bono responsabilidad", "responsabilidad"], after_idx=idx_last_agrupacion)
        idx_bono_gest = get_col_idx(["bono gestion variable", "bono gestion", "bono gestin variable", "gestion"], after_idx=idx_last_agrupacion)
        idx_bono_perm = get_col_idx(["bono de permanencia", "bono permanencia", "permanencia"], after_idx=idx_last_agrupacion)
        
        idx_grat = get_col_idx(["gratificacion", "gratificacin"])
        idx_col = get_col_idx(["colacion", "colacin"])
        idx_mov = get_col_idx(["movilizacion", "movilizacion"])
        idx_pasajes = get_col_idx(["pasajes"], after_idx=idx_last_agrupacion)
        idx_traslados = get_col_idx(["traslados"], after_idx=idx_last_agrupacion)
        idx_estudios = get_col_idx(["bono estudios", "estudios"], after_idx=idx_last_agrupacion)
        idx_fallecimiento = get_col_idx(["bono fallecimiento", "fallecimiento"], after_idx=idx_last_agrupacion)
        
        idx_apvi = get_col_idx(["apvi", "apv"], after_idx=idx_last_agrupacion)
        idx_anticipo = get_col_idx(["anticipo"], after_idx=idx_last_agrupacion)
        idx_ccaf_credito = get_col_idx(["ccaf credito", "credito ccaf", "credito caja", "credito araucana", "cajacred", "creditos personales ccaf"], after_idx=idx_last_agrupacion)
        idx_ccaf_prestamo = get_col_idx(["ccaf prestamo", "prestamo ccaf", "prestamo caja comp. la araucana", "prestamo caja", "la araucana", "araucana", "prestamo caja la araucana", "prestamocaraucana", "prestamo c araucana", "los heroes", "heroes", "caja los heroes", "18 de septiembre", "caja 18", "los andes", "caja los andes"], after_idx=idx_last_agrupacion)
        idx_retencion_judicial = get_col_idx(["retencion judicial"], after_idx=idx_last_agrupacion)
        idx_prestamos_empresa = get_col_idx(["prestamos empresa", "prestamo empresa"], after_idx=idx_last_agrupacion)
        idx_seguro_complementario = get_col_idx(["seguro complementario"], after_idx=idx_last_agrupacion)
        idx_falp = get_col_idx(["falp"], after_idx=idx_last_agrupacion)
        idx_ahorro_afp = get_col_idx(["ahorro afp"], after_idx=idx_last_agrupacion)
        idx_licencia_dias = get_col_idx(["dias con licencia medica", "licenciadias", "licencia"])
        idx_periodo = get_col_idx(["proceso"])

        # Dynamic mapping for exponibles/non-exponibles
        earning_mappings = {}
        unmapped_indices = set()
        if True:
            explicit_idxs = {idx_rut, idx_contrato, idx_dias, idx_sueldo_base, idx_grat, idx_col, idx_mov, idx_pasajes, idx_traslados, idx_estudios, idx_fallecimiento, idx_apvi, idx_anticipo, idx_ccaf_credito, idx_ccaf_prestamo, idx_retencion_judicial, idx_prestamos_empresa, idx_seguro_complementario, idx_falp, idx_ahorro_afp, idx_licencia_dias, idx_periodo}
            
            start_mapping_idx = idx_sueldo_base + 1 if idx_sueldo_base is not None else 1
                
            for idx in range(start_mapping_idx, len(headers_row)):
                if idx in explicit_idxs:
                    continue
                header = headers_row[idx]
                if not header:
                    continue
                h_clean = clean(header)
                
                # Skip metadata and non-payment columns
                skip_terms = ["justificacion", "observacion", "observaciones", "sindicato", "pauta", "banco", "cuenta", "turno", "cargo", "inicio", "termino", "nombre", "empresa", "proceso", "total", "afecto", "liquido", "costo", "haberes", "descuento", "rebaja", "afp", "salud", "isapre", "contrato", "rut", "dias", "licencia", "ccaf", "mutual", "seguro", "aporte", "fapp", "sis", "impuesto", "cc", "apellido"]
                if any(term in h_clean for term in skip_terms):
                    continue
                
                is_severance = False
                for term in ["ias", "vacaciones", "indemnizacion", "indemnizacin", "tiempo servido", "aviso", "finiquito", "licencia", "permiso"]:
                    if term in h_clean:
                        if term == "ias" and "dias" in h_clean:
                            continue
                        is_severance = True
                        break
                if is_severance:
                    continue
                
                is_non_imp = False
                for term in ["colacion", "colacin", "movilizacion", "movilizacin", "pasajes", "traslados", "viatico", "cargas", "sala cuna", "estudios", "fallecimiento", "sobregiro", "devolucion", "reembolso", "ccaf", "familiar", "retroactiva"]:
                    if term in h_clean:
                        is_non_imp = True
                        break
                        
                if is_non_imp:
                    if "colacion" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "colacion")
                    elif "movilizacion" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "movilizacion")
                    elif "pasajes" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "pasajes")
                    elif "traslados" in h_clean or "viatico" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "traslados")
                    elif "fallecimiento" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "fallecimiento")
                    elif "estudios" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "bono_estudios")
                    else:
                        unmapped_indices.add(idx)
                else:
                    if "diferencia" in h_clean and ("gratificacion" in h_clean or "gratificacin" in h_clean):
                        earning_mappings[idx] = ("imponible", "diferencia_gratificacion")
                        continue
                    if h_clean == "gratificacion" or h_clean == "gratificacin":
                        continue
                    if "descanso" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_descanso")
                    elif "feriado" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_feriado")
                    elif "responsabilidad" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_responsabilidad")
                    elif "gestion" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_gestion")
                    elif "permanencia" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_permanencia")
                    elif "incentivo" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_incentivo")
                    else:
                        unmapped_indices.add(idx)

        count = 0
        consecutive_empty = 0
        for row in sheet.iter_rows(min_row=header_row_idx+1, values_only=True):
            safe_rut_idx = idx_rut if idx_rut is not None else 1
            safe_contrato_idx = idx_contrato if idx_contrato is not None else 0
            if not row or len(row) <= max(safe_rut_idx, safe_contrato_idx) or not row[safe_rut_idx]:
                consecutive_empty += 1
                if consecutive_empty > 50:
                    break
                continue
            consecutive_empty = 0

            rut = clean_rut(row[safe_rut_idx])
            contrato = int(row[idx_contrato]) if idx_contrato is not None and row[idx_contrato] is not None else 1
            dias_trabajados = int(row[idx_dias]) if idx_dias is not None and row[idx_dias] is not None else 30
            sueldo_base = row[idx_sueldo_base] if idx_sueldo_base is not None and row[idx_sueldo_base] is not None else 0.0
            
            # Initialize earnings
            bono_desc = 0.0
            bono_feri = 0.0
            bono_resp = 0.0
            bono_gest = 0.0
            bono_perm = 0.0
            bono_inc = 0.0
            diferencia_gratificacion = 0.0
            colacion = float(row[idx_col]) if idx_col is not None and idx_col < len(row) and row[idx_col] is not None else 0.0
            movilizacion = float(row[idx_mov]) if idx_mov is not None and idx_mov < len(row) and row[idx_mov] is not None else 0.0
            pasajes = float(row[idx_pasajes]) if idx_pasajes is not None and idx_pasajes < len(row) and row[idx_pasajes] is not None else 0.0
            traslados = float(row[idx_traslados]) if idx_traslados is not None and idx_traslados < len(row) and row[idx_traslados] is not None else 0.0
            bono_fallecimiento = float(row[idx_fallecimiento]) if idx_fallecimiento is not None and idx_fallecimiento < len(row) and row[idx_fallecimiento] is not None else 0.0
            bono_estudios = float(row[idx_estudios]) if idx_estudios is not None and idx_estudios < len(row) and row[idx_estudios] is not None else 0.0
            
            # Audit for duplicates
            conceptos_vistos = {}
            for b_idx, (earn_type, target) in earning_mappings.items():
                if b_idx < len(row) and row[b_idx] is not None:
                    try:
                        val = float(row[b_idx])
                    except:
                        val = 0.0
                    
                    if val > 0:
                        header = headers_row[b_idx]
                        if header:
                            h_clean = target 
                            if h_clean in conceptos_vistos:
                                prev_header, prev_val = conceptos_vistos[h_clean]
                                period_val = str(row[idx_periodo]).strip() if idx_periodo is not None and idx_periodo < len(row) and row[idx_periodo] is not None else ""
                                detail_msg = f"Concepto '{header}' (mapeado a '{target}') duplicado con montos ${prev_val:,.0f} y ${val:,.0f}"
                                cursor.execute("""
                                    INSERT OR IGNORE INTO alertas_auditoria (rut, contrato, periodo, tipo_alerta, detalle) VALUES (?, ?, ?, ?, ?)
                                """, (rut, contrato, period_val, "Concepto Duplicado", detail_msg))
                            else:
                                conceptos_vistos[h_clean] = (header, val)
                    
                    if earn_type == "imponible":
                        if target == "bono_descanso": bono_desc += val
                        elif target == "bono_feriado": bono_feri += val
                        elif target == "bono_responsabilidad": bono_resp += val
                        elif target == "bono_gestion": bono_gest += val
                        elif target == "bono_permanencia": bono_perm += val
                        elif target == "bono_incentivo": bono_inc += val
                        elif target == "diferencia_gratificacion": diferencia_gratificacion += val
                    elif earn_type == "non_imponible":
                        if target == "colacion": colacion += val
                        elif target == "movilizacion": movilizacion += val
                        elif target == "pasajes": pasajes += val
                        elif target == "traslados": traslados += val
                        elif target == "bono_estudios": bono_estudios += val
                        elif target == "bono_fallecimiento": bono_fallecimiento += val

            # Audit for unmapped columns with values > 0
            for u_idx in unmapped_indices:
                if u_idx < len(row) and row[u_idx] is not None:
                    try:
                        val = float(row[u_idx])
                    except:
                        val = 0.0
                    
                    if val > 0:
                        header = headers_row[u_idx]
                        period_val = str(row[idx_periodo]).strip() if idx_periodo is not None and idx_periodo < len(row) and row[idx_periodo] is not None else ""
                        if not period_val:
                            # Deduce period from file name if missing
                            fname_lower = os.path.basename(filepath).lower()
                            if "enero" in fname_lower: period_val = "2026-01"
                            elif "febrero" in fname_lower: period_val = "2026-02"
                            elif "marzo" in fname_lower: period_val = "2026-03"
                            elif "abril" in fname_lower: period_val = "2026-04"
                            elif "mayo" in fname_lower: period_val = "2026-05"
                            elif "junio" in fname_lower: period_val = "2026-06"
                            else:
                                for key_p in ['2026-01', '2026-02', '2026-03', '2026-04', '2026-05', '2026-06']:
                                    if key_p in fname_lower:
                                        period_val = key_p
                                        break
                            if not period_val:
                                period_val = "2026-06" if "junio" in fname_lower else "2026-05"
                                
                        detail_msg = f"Concepto no mapeado '{header}' detectado con monto ${val:,.0f}."
                        cursor.execute("""
                            INSERT OR IGNORE INTO alertas_auditoria (rut, contrato, periodo, tipo_alerta, detalle) VALUES (?, ?, ?, ?, ?)
                        """, (rut, contrato, period_val, "Concepto No Mapeado", detail_msg))

            gratificacion = float(row[idx_grat]) if idx_grat is not None and row[idx_grat] is not None else 0.0
            apvi = float(row[idx_apvi]) if idx_apvi is not None and row[idx_apvi] is not None else 0.0
            
            periodo = str(row[idx_periodo]).strip() if idx_periodo is not None and row[idx_periodo] is not None else ""
            if not periodo:
                # Deduce period from file name if missing
                fname_lower = os.path.basename(filepath).lower()
                if "enero" in fname_lower: periodo = "2026-01"
                elif "febrero" in fname_lower: periodo = "2026-02"
                elif "marzo" in fname_lower: periodo = "2026-03"
                elif "abril" in fname_lower: periodo = "2026-04"
                elif "mayo" in fname_lower: periodo = "2026-05"
                elif "junio" in fname_lower: periodo = "2026-06"
                else:
                    for key_p in ['2026-01', '2026-02', '2026-03', '2026-04', '2026-05', '2026-06']:
                        if key_p in fname_lower:
                            periodo = key_p
                            break
                if not periodo:
                    periodo = "2026-06" if "junio" in fname_lower else "2026-05"

            licencia_dias = float(row[idx_licencia_dias]) if idx_licencia_dias is not None and idx_licencia_dias < len(row) and row[idx_licencia_dias] is not None else 0.0

            anticipo_val = float(row[idx_anticipo]) if idx_anticipo is not None and idx_anticipo < len(row) and row[idx_anticipo] is not None else 0.0
            ccaf_credito_val = float(row[idx_ccaf_credito]) if idx_ccaf_credito is not None and idx_ccaf_credito < len(row) and row[idx_ccaf_credito] is not None else 0.0
            ccaf_prestamo_val = float(row[idx_ccaf_prestamo]) if idx_ccaf_prestamo is not None and idx_ccaf_prestamo < len(row) and row[idx_ccaf_prestamo] is not None else 0.0
            retencion_judicial_val = float(row[idx_retencion_judicial]) if idx_retencion_judicial is not None and idx_retencion_judicial < len(row) and row[idx_retencion_judicial] is not None else 0.0
            prestamos_empresa_val = float(row[idx_prestamos_empresa]) if idx_prestamos_empresa is not None and idx_prestamos_empresa < len(row) and row[idx_prestamos_empresa] is not None else 0.0
            seguro_complementario_val = float(row[idx_seguro_complementario]) if idx_seguro_complementario is not None and idx_seguro_complementario < len(row) and row[idx_seguro_complementario] is not None else 0.0
            falp_val = float(row[idx_falp]) if idx_falp is not None and idx_falp < len(row) and row[idx_falp] is not None else 0.0
            ahorro_afp_val = float(row[idx_ahorro_afp]) if idx_ahorro_afp is not None and idx_ahorro_afp < len(row) and row[idx_ahorro_afp] is not None else 0.0

            obs_val = ""
            # Load observations if present
            if len(row) > 100 and row[100]:
                obs_val = str(row[100]).strip()
            elif len(row) > 99 and row[99]:
                obs_val = str(row[99]).strip()

            just_json_str = "{}"
            period_extra = all_extra_data.get(periodo, {})
            if period_extra:
                extra = period_extra.get((rut, contrato), {})
                if extra:
                    justs = extra.get("justificaciones", {}).copy()
                    just_json_str = json.dumps(justs, ensure_ascii=False)

            cursor.execute("""
            INSERT OR REPLACE INTO planilla_entradas (
                rut, contrato, periodo, dias_trabajados, sueldo_base, bono_descanso, bono_feriado,
                bono_incentivo, bono_responsabilidad, bono_gestion, bono_permanencia, gratificacion, diferencia_gratificacion,
                colacion, movilizacion, pasajes, traslados, bono_estudios, bono_fallecimiento,
                apvi, anticipo, ccaf_credito, ccaf_prestamo, retencion_judicial, prestamos_empresa, seguro_complementario, falp, ahorro_afp,
                licencia_dias, justificaciones_json, observaciones, hash_origen, timestamp_carga
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rut, contrato, periodo, dias_trabajados, sueldo_base, bono_desc, bono_feri,
                bono_inc, bono_resp, bono_gest, bono_perm, gratificacion, diferencia_gratificacion,
                colacion, movilizacion, pasajes, traslados, bono_estudios, bono_fallecimiento,
                apvi, anticipo_val, ccaf_credito_val, ccaf_prestamo_val, retencion_judicial_val, prestamos_empresa_val, seguro_complementario_val, falp_val, ahorro_afp_val,
                licencia_dias, just_json_str, obs_val, file_hash, timestamp_now
            ))
            count += 1
        total_count += count
        print(f"Loaded {count} inputs from {os.path.basename(filepath)}")

    conn.commit()
    conn.close()
    print(f"Total inputs loaded into planilla_entradas: {total_count}")

def load_rex_comparisons(custom_files=None):
    print("Loading Rex+ payroll comparisons...")
    if custom_files:
        files = custom_files
    else:
        files = glob.glob(os.path.join(BASE_DIR, "Detalle process *.xlsx"))
        # Fallback to general Detalle proceso *.xlsx glob
        if not files:
            files = glob.glob(os.path.join(BASE_DIR, "Detalle proceso *.xlsx"))
    
    # Sort files so that files with '(' (specific processes) are loaded after the base files
    files = sorted(files, key=lambda x: (1 if '(' in os.path.basename(x) else 0, x))
    if not files:
        print("No Rex+ detail files found matching 'Detalle proceso *.xlsx' or custom files.")
        return
    
    # Load all justifications (May and historical)
    all_extra_data = load_all_justifications()

    conn = get_connection()
    cursor = conn.cursor()
    
    # Ensure date columns exist in rex_comparisons
    try:
        cursor.execute("ALTER TABLE rex_comparisons ADD COLUMN fecha_inicio TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE rex_comparisons ADD COLUMN fecha_termino TEXT")
    except sqlite3.OperationalError:
        pass
        
    total_count = 0

    for filepath in files:
        print(f"Parsing: {os.path.basename(filepath)}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Determine the sheet name: 'Detalle' or monthly sheets ('06-2026', '05-2026', etc.)
        sheetname = None
        for s in ['Detalle', '06-2026', '05-2026', '04-2026', '03-2026', '02-2026', '01-2026', 'Sheet1']:
            if s in wb.sheetnames:
                sheetname = s
                break
        if not sheetname:
            print(f"Skipping {filepath} - no valid sheet found ('Detalle' or monthly sheet).")
            continue
            
        sheet = wb[sheetname]

        # Row 5 or 6 has headers. Let's find it dynamically
        headers_row = None
        data_start_row = 6
        for r_idx in [5, 6]:
            row_vals = list(sheet.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            if any(isinstance(val, str) and 'rut' in val.lower() for val in row_vals if val):
                headers_row = row_vals
                data_start_row = r_idx + 1
                break
        if headers_row is None:
            # Fallback
            headers_row = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))[0]
            data_start_row = 6
        
        def clean(s):
            if not s:
                return ""
            s = str(s).lower().strip()
            s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
            s = s.replace("ñ", "n").replace("ü", "u")
            s = s.replace("  ", " ")
            return s

        def get_col_idx(candidates):
            cleaned_candidates = [clean(c) for c in candidates]
            for idx, h in enumerate(headers_row):
                if h and clean(h) in cleaned_candidates:
                    return idx
            for idx, h in enumerate(headers_row):
                if h:
                    ch = clean(h)
                    for cc in cleaned_candidates:
                        if cc and ch and (cc in ch or ch in cc):
                            return idx
            return None

        idx_rut = get_col_idx(["rut"])
        idx_contrato = get_col_idx(["contrato"])
        idx_nombre = get_col_idx(["nombre"])
        idx_dias = get_col_idx(["dias trabajados", "das trabajados", "trabajados"])
        idx_sueldo_base = get_col_idx(["sueldo base"])
        
        idx_bono_desc = get_col_idx(["bono descanso", "descanso"])
        idx_bono_feri = get_col_idx(["bono feriado", "feriado"])
        idx_bono_inc = get_col_idx(["bono incentivo variable", "bono incentivo", "incentivo"])
        idx_bono_resp = get_col_idx(["bono responsabilidad", "responsabilidad"])
        idx_bono_gest = get_col_idx(["bono gestion variable", "bono gestion", "bono gestin variable", "gestion"])
        idx_bono_perm = get_col_idx(["bono de permanencia", "bono permanencia", "permanencia"])
        
        idx_grat = get_col_idx(["gratificacion", "gratificacin"])
        idx_col = get_col_idx(["colacion", "colacin"])
        idx_mov = get_col_idx(["movilizacion", "movilizacin"])
        idx_pasajes = get_col_idx(["pasajes"])
        idx_traslados = get_col_idx(["traslados"])
        idx_estudios = get_col_idx(["bono estudios", "estudios"])
        idx_fallecimiento = get_col_idx(["bono fallecimiento", "fallecimiento"])
        
        idx_apvi = get_col_idx(["apvi ahorro voluntario mensual", "apvi", "ahorro voluntario"])
        idx_afp_cotiz = get_col_idx(["cotizacion afp", "cotizacin afp"])
        idx_salud_cotiz = get_col_idx(["cotizacion salud", "cotizacin salud"])
        idx_cesantia_trab = get_col_idx(["seguro de cesantia", "seguro de cesanta", "seguro cesantia"])
        idx_impuesto = get_col_idx(["impuesto", "impuesto unico", "impuesto nico"])
        idx_total_desc = get_col_idx(["total descuentos"])
        idx_sueldo_liq = get_col_idx(["sueldo liquido", "sueldo lquido"])
        idx_alcance_liq = get_col_idx(["alcance liquido", "alcance lquido"])
        
        idx_mutual = get_col_idx(["mutual", "mutualidad"])
        idx_sis = get_col_idx(["seguro invalidez y sobrevivencia", "sis"])

        idx_anticipo = get_col_idx(["anticipo"])
        idx_ccaf_credito = get_col_idx(["creditos personales ccaf", "creditos ccaf"])
        idx_ccaf_prestamo = get_col_idx(["prestamo caja comp", "prestamo araucana"])
        idx_retencion_judicial = get_col_idx(["retencion judicial"])
        idx_prestamos_empresa = get_col_idx(["prestamos empresa"])
        idx_seguro_complementario = get_col_idx(["seguro complementario"])
        idx_falp = get_col_idx(["falp"])
        
        idx_afc_ci = get_col_idx(["seguro de cesantia ci", "cesantia ci"])
        idx_afc_solid = get_col_idx(["seguro de cesantia solidario", "cesantia solidario"])
        
        idx_costo = get_col_idx(["costo empresa"])
        idx_total_imp = get_col_idx(["total imponible sin tope", "total imponible", "imponible"])
        idx_afecto_afp = get_col_idx(["afecto afp"])
        idx_afecto_ces = get_col_idx(["afecto cesantia", "afecto cesanta"])
        idx_afecto_imp = get_col_idx(["afecto impuesto"])
        
        idx_afp_name = get_col_idx(["afp"])
        idx_salud_name = get_col_idx(["inst. salud", "salud"])
        idx_tipo_contrato = get_col_idx(["tipo contrato"])
        idx_periodo = get_col_idx(["proceso"])
        idx_fecha_inicio = get_col_idx(["fecha inicio", "fecha inicio contrato", "fecha ingreso"])
        idx_fecha_termino = get_col_idx(["fecha termino", "fecha termino contrato", "fecha retiro"])

        # HR and Severance Columns
        idx_licencia_dias = get_col_idx(["dias con licencia medica", "licenciadias", "licencia"])
        idx_ias_vacaciones = get_col_idx(["ias vacaciones legales", "iasvacaciones"])
        idx_ias_anos_servicio_1 = get_col_idx(["indemnizacion de tiempo servido", "indemnizacion tiempo servido", "indemnización de tiempo servido"])
        idx_ias_anos_servicio_2 = get_col_idx(["indemnizacion legal", "indemnización legal"])
        idx_ias_aviso = get_col_idx(["ias mes de aviso", "mes de aviso"])
        idx_ias_anticipo_finiquito = get_col_idx(["anticipo de finiquito", "anticipofiniquito"])
        idx_ias_seguro_cesantia = get_col_idx(["seguro cesantia recuperado", "seguro cesantía recuperado"])

        # Role and Location columns for historical mappings
        idx_sede = get_col_idx(["sede"])
        idx_cargo = get_col_idx(["cargo"])
        idx_cc = get_col_idx(["centro costo", "centro_costo"])

        # Build dynamic earning mappings to classify each index in the earning columns range
        first_prev_idx = min(idx for idx in [idx_apvi, idx_afp_cotiz, idx_salud_cotiz, idx_cesantia_trab] if idx is not None)
        earning_mappings = {}
        if idx_sueldo_base is not None and first_prev_idx is not None:
            for idx in range(idx_sueldo_base + 1, first_prev_idx):
                header = headers_row[idx]
                if not header:
                    continue
                h_clean = clean(header)
                
                # Check if it is a severance/HR column to skip (parsed separately)
                is_severance = False
                for term in ["ias", "vacaciones", "indemnizacion", "indemnizacin", "tiempo servido", "aviso", "finiquito", "licencia", "permiso"]:
                    if term in h_clean:
                        if term == "ias" and "dias" in h_clean:
                            continue
                        is_severance = True
                        break
                if is_severance:
                    continue
                
                # Check if it is a known non-imponible
                is_non_imp = False
                for term in ["colacion", "colacin", "movilizacion", "movilizacin", "pasajes", "traslados", "viatico", "cargas", "sala cuna", "estudios", "fallecimiento", "sobregiro", "devolucion", "reembolso", "ccaf", "familiar", "retroactiva"]:
                    if term in h_clean:
                        is_non_imp = True
                        break
                        
                if is_non_imp:
                    # Classify into specific non-imponibles
                    if "colacion" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "colacion")
                    elif "movilizacion" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "movilizacion")
                    elif "pasajes" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "pasajes")
                    elif "traslados" in h_clean or "viatico" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "traslados")
                    elif "fallecimiento" in h_clean:
                        earning_mappings[idx] = ("non_imponible", "fallecimiento")
                    else:
                        # e.g. bono estudios, sala cuna sum to bono_estudios
                        earning_mappings[idx] = ("non_imponible", "bono_estudios")
                else:
                    # It is imponible! (unless it's the gratificacion itself, which we parse separately)
                    if "diferencia" in h_clean and ("gratificacion" in h_clean or "gratificacin" in h_clean):
                        earning_mappings[idx] = ("imponible", "diferencia_gratificacion")
                        continue
                    if h_clean == "gratificacion" or h_clean == "gratificacin":
                        continue
                    # Classify into specific imponibles
                    if "descanso" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_descanso")
                    elif "feriado" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_feriado")
                    elif "responsabilidad" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_responsabilidad")
                    elif "gestion" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_gestion")
                    elif "permanencia" in h_clean:
                        earning_mappings[idx] = ("imponible", "bono_permanencia")
                    else:
                        # Others sum to bono_incentivo
                        earning_mappings[idx] = ("imponible", "bono_incentivo")

        count = 0
        # Data starts from Row data_start_row
        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            if not row or idx_rut is None or len(row) <= idx_rut or not row[idx_rut]:
                continue

            rut = clean_rut(row[idx_rut])
            contrato = int(row[idx_contrato]) if idx_contrato is not None and idx_contrato < len(row) and row[idx_contrato] is not None else 1
            nombre = row[idx_nombre] if idx_nombre is not None and idx_nombre < len(row) else ""
            dias_trabajados = int(row[idx_dias]) if idx_dias is not None and idx_dias < len(row) and row[idx_dias] is not None else 30
            sueldo_base = float(row[idx_sueldo_base]) if idx_sueldo_base is not None and idx_sueldo_base < len(row) and row[idx_sueldo_base] is not None else 0.0
            
            # Initialize earnings
            bono_desc = 0.0
            bono_feri = 0.0
            bono_resp = 0.0
            bono_gest = 0.0
            bono_perm = 0.0
            bono_inc = 0.0
            diferencia_gratificacion = 0.0
            colacion = 0.0
            movilizacion = 0.0
            pasajes = 0.0
            traslados = 0.0
            bono_fallecimiento = 0.0
            bono_estudios = 0.0
            
            # Accumulate dynamically mapped earnings and audit for duplicates
            conceptos_vistos = {}
            for b_idx, (earn_type, target) in earning_mappings.items():
                if b_idx < len(row) and row[b_idx] is not None:
                    try:
                        val = float(row[b_idx])
                    except:
                        val = 0.0
                    
                    if val > 0:
                        header = headers_row[b_idx]
                        if header:
                            h_clean = clean(header)
                            if h_clean in conceptos_vistos:
                                prev_header, prev_val = conceptos_vistos[h_clean]
                                period_val = str(row[idx_periodo]).strip() if idx_periodo is not None and idx_periodo < len(row) and row[idx_periodo] is not None else ""
                                detail_msg = f"Concepto '{header}' duplicado con montos ${prev_val:,.0f} y ${val:,.0f}"
                                cursor.execute(
                                    "                                    INSERT OR IGNORE INTO alertas_auditoria (rut, contrato, periodo, tipo_alerta, detalle) VALUES (?, ?, ?, ?, ?)",
                                    (rut, contrato, period_val, "Concepto Duplicado", detail_msg)
                                )
                            else:
                                conceptos_vistos[h_clean] = (header, val)
                    
                    if earn_type == "imponible":
                        if target == "bono_descanso": bono_desc += val
                        elif target == "bono_feriado": bono_feri += val
                        elif target == "bono_responsabilidad": bono_resp += val
                        elif target == "bono_gestion": bono_gest += val
                        elif target == "bono_permanencia": bono_perm += val
                        elif target == "bono_incentivo": bono_inc += val
                        elif target == "diferencia_gratificacion": diferencia_gratificacion += val
                    elif earn_type == "non_imponible":
                        if target == "colacion": colacion += val
                        elif target == "movilizacion": movilizacion += val
                        elif target == "pasajes": pasajes += val
                        elif target == "traslados": traslados += val
                        elif target == "bono_fallecimiento": bono_fallecimiento += val
                        elif target == "bono_estudios": bono_estudios += val

            gratificacion = row[idx_grat] if idx_grat is not None and row[idx_grat] is not None else 0.0
            
            apvi = row[idx_apvi] if idx_apvi is not None and row[idx_apvi] is not None else 0.0
            cotizacion_afp = row[idx_afp_cotiz] if idx_afp_cotiz is not None and row[idx_afp_cotiz] is not None else 0.0
            cotizacion_salud = row[idx_salud_cotiz] if idx_salud_cotiz is not None and row[idx_salud_cotiz] is not None else 0.0
            seguro_cesantia_trab = row[idx_cesantia_trab] if idx_cesantia_trab is not None and row[idx_cesantia_trab] is not None else 0.0
            impuesto = row[idx_impuesto] if idx_impuesto is not None and row[idx_impuesto] is not None else 0.0
            total_descuentos = row[idx_total_desc] if idx_total_desc is not None and row[idx_total_desc] is not None else 0.0
            sueldo_liquido = row[idx_sueldo_liq] if idx_sueldo_liq is not None and row[idx_sueldo_liq] is not None else 0.0
            alcance_liquido = row[idx_alcance_liq] if idx_alcance_liq is not None and row[idx_alcance_liq] is not None else 0.0
            
            mutual = row[idx_mutual] if idx_mutual is not None and row[idx_mutual] is not None else 0.0
            sis = row[idx_sis] if idx_sis is not None and row[idx_sis] is not None else 0.0
            
            afc_ci = row[idx_afc_ci] if idx_afc_ci is not None and row[idx_afc_ci] is not None else 0.0
            afc_solid = row[idx_afc_solid] if idx_afc_solid is not None and row[idx_afc_solid] is not None else 0.0
            seguro_cesantia_emp = afc_ci + afc_solid
            
            costo_empresa = row[idx_costo] if idx_costo is not None and row[idx_costo] is not None else 0.0
            total_imponible = row[idx_total_imp] if idx_total_imp is not None and row[idx_total_imp] is not None else 0.0
            afecto_afp = row[idx_afecto_afp] if idx_afecto_afp is not None and row[idx_afecto_afp] is not None else 0.0
            afecto_cesantia = row[idx_afecto_ces] if idx_afecto_ces is not None and row[idx_afecto_ces] is not None else 0.0
            afecto_impuesto = row[idx_afecto_imp] if idx_afecto_imp is not None and row[idx_afecto_imp] is not None else 0.0
            
            afp_name = row[idx_afp_name] if idx_afp_name is not None else ""
            salud_name = row[idx_salud_name] if idx_salud_name is not None else ""
            tipo_contrato = row[idx_tipo_contrato] if idx_tipo_contrato is not None else ""
            periodo = str(row[idx_periodo]).strip() if idx_periodo is not None and row[idx_periodo] is not None else ""
            fecha_inicio_val = format_date(row[idx_fecha_inicio]) if idx_fecha_inicio is not None and idx_fecha_inicio < len(row) and row[idx_fecha_inicio] is not None else None
            fecha_termino_val = format_date(row[idx_fecha_termino]) if idx_fecha_termino is not None and idx_fecha_termino < len(row) and row[idx_fecha_termino] is not None else None

            # HR and Severances values
            licencia_dias = float(row[idx_licencia_dias]) if idx_licencia_dias is not None and idx_licencia_dias < len(row) and row[idx_licencia_dias] is not None else 0.0
            ias_vacaciones = float(row[idx_ias_vacaciones]) if idx_ias_vacaciones is not None and idx_ias_vacaciones < len(row) and row[idx_ias_vacaciones] is not None else 0.0
            ias_anos_servicio_1 = float(row[idx_ias_anos_servicio_1]) if idx_ias_anos_servicio_1 is not None and idx_ias_anos_servicio_1 < len(row) and row[idx_ias_anos_servicio_1] is not None else 0.0
            ias_anos_servicio_2 = float(row[idx_ias_anos_servicio_2]) if idx_ias_anos_servicio_2 is not None and idx_ias_anos_servicio_2 < len(row) and row[idx_ias_anos_servicio_2] is not None else 0.0
            ias_anos_servicio = ias_anos_servicio_1 + ias_anos_servicio_2
            ias_aviso = float(row[idx_ias_aviso]) if idx_ias_aviso is not None and idx_ias_aviso < len(row) and row[idx_ias_aviso] is not None else 0.0
            ias_anticipo_finiquito = float(row[idx_ias_anticipo_finiquito]) if idx_ias_anticipo_finiquito is not None and idx_ias_anticipo_finiquito < len(row) and row[idx_ias_anticipo_finiquito] is not None else 0.0
            ias_seguro_cesantia_recuperado = float(row[idx_ias_seguro_cesantia]) if idx_ias_seguro_cesantia is not None and idx_ias_seguro_cesantia < len(row) and row[idx_ias_seguro_cesantia] is not None else 0.0


            # Sede, Cargo, CC values
            sede_val = row[idx_sede] if idx_sede is not None and idx_sede < len(row) and row[idx_sede] is not None else ""
            cargo_val = str(row[idx_cargo]).strip().upper() if idx_cargo is not None and idx_cargo < len(row) and row[idx_cargo] is not None else ""
            cc_val = row[idx_cc] if idx_cc is not None and idx_cc < len(row) and row[idx_cc] is not None else ""

            anticipo_val = float(row[idx_anticipo]) if idx_anticipo is not None and idx_anticipo < len(row) and row[idx_anticipo] is not None else 0.0
            ccaf_credito_val = float(row[idx_ccaf_credito]) if idx_ccaf_credito is not None and idx_ccaf_credito < len(row) and row[idx_ccaf_credito] is not None else 0.0
            ccaf_prestamo_val = float(row[idx_ccaf_prestamo]) if idx_ccaf_prestamo is not None and idx_ccaf_prestamo < len(row) and row[idx_ccaf_prestamo] is not None else 0.0
            retencion_judicial_val = float(row[idx_retencion_judicial]) if idx_retencion_judicial is not None and idx_retencion_judicial < len(row) and row[idx_retencion_judicial] is not None else 0.0
            prestamos_empresa_val = float(row[idx_prestamos_empresa]) if idx_prestamos_empresa is not None and idx_prestamos_empresa < len(row) and row[idx_prestamos_empresa] is not None else 0.0
            seguro_complementario_val = float(row[idx_seguro_complementario]) if idx_seguro_complementario is not None and idx_seguro_complementario < len(row) and row[idx_seguro_complementario] is not None else 0.0
            falp_val = float(row[idx_falp]) if idx_falp is not None and idx_falp < len(row) and row[idx_falp] is not None else 0.0

            # Override for all justifications
            lic_dias_val = licencia_dias
            just_json_str = "{}"
            
            period_extra = all_extra_data.get(periodo, {})
            if period_extra:
                extra = period_extra.get((rut, contrato), {})
                if extra:
                    justs = extra.get("justificaciones", {}).copy()
                    if extra.get("observaciones"):
                        justs["Observaciones Generales"] = extra["observaciones"]
                    just_json_str = json.dumps(justs, ensure_ascii=False)

            cursor.execute("""
            INSERT OR REPLACE INTO rex_comparisons (
                rut, contrato, nombre, dias_trabajados, sueldo_base, bono_descanso, bono_feriado,
                bono_incentivo, bono_responsabilidad, bono_gestion, bono_permanencia, gratificacion,
                colacion, movilizacion, pasajes, traslados, bono_estudios, bono_fallecimiento,
                apvi, cotizacion_afp, cotizacion_salud, seguro_cesantia_trab, impuesto, total_descuentos,
                sueldo_liquido, alcance_liquido, mutual, sis, seguro_cesantia_emp, costo_empresa, total_imponible,
                afecto_afp, afecto_cesantia, afecto_impuesto, afp, isapre, tipo_contrato, periodo,
                licencia_dias, ias_vacaciones, ias_anos_servicio, ias_aviso, justificaciones_json,
                cargo, centro_costo, sede, fecha_inicio, fecha_termino,
                anticipo, ccaf_credito, ccaf_prestamo, retencion_judicial, prestamos_empresa, seguro_complementario, falp
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rut, contrato, nombre, dias_trabajados, sueldo_base, bono_desc, bono_feri,
                bono_inc, bono_resp, bono_gest, bono_perm, gratificacion,
                colacion, movilizacion, pasajes, traslados, bono_estudios, bono_fallecimiento,
                apvi, cotizacion_afp, cotizacion_salud, seguro_cesantia_trab, impuesto, total_descuentos,
                sueldo_liquido, alcance_liquido, mutual, sis, seguro_cesantia_emp, costo_empresa, total_imponible,
                afecto_afp, afecto_cesantia, afecto_impuesto, afp_name, salud_name, tipo_contrato, periodo,
                lic_dias_val, ias_vacaciones, ias_anos_servicio, ias_aviso, just_json_str,
                cargo_val, cc_val, sede_val, fecha_inicio_val, fecha_termino_val,
                anticipo_val, ccaf_credito_val, ccaf_prestamo_val, retencion_judicial_val, prestamos_empresa_val, seguro_complementario_val, falp_val
            ))
            count += 1
        total_count += count
        print(f"Loaded {count} comparisons for period {periodo}")

    conn.commit()
    conn.close()
    print(f"Loaded total of {total_count} Rex+ payroll comparison records.")
    fill_missing_employees()


def fill_missing_employees():
    print("Populating missing historical employees in database...")
    conn = get_connection()
    cursor = conn.cursor()
    
    # Ensure new columns exist
    try:
        cursor.execute("ALTER TABLE empleados ADD COLUMN agrupacion TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE empleados ADD COLUMN area TEXT")
    except sqlite3.OperationalError:
        pass

    cursor.execute("""
        SELECT DISTINCT c.rut, c.contrato, c.nombre, c.afp, c.isapre, c.tipo_contrato, 
                        c.sueldo_base, c.cotizacion_salud, c.periodo,
                        c.cargo, c.centro_costo, c.sede,
                        c.fecha_inicio, c.fecha_termino, c.dias_trabajados
        FROM rex_comparisons c
        LEFT JOIN empleados e ON c.rut = e.rut AND c.contrato = e.contrato
        WHERE e.rut IS NULL
    """)
    missing = cursor.fetchall()
    print(f"Found {len(missing)} missing historical contract records to populate.")
    
    # Build centro_costo to id_obra mapping from existing records to resolve it for historical ones
    cursor.execute("SELECT DISTINCT centro_costo, id_obra FROM empleados WHERE id_obra IS NOT NULL AND id_obra != ''")
    cc_to_id_obra = {row[0]: row[1] for row in cursor.fetchall() if row[0]}
    
    for r in missing:
        rut = r["rut"]
        contrato = r["contrato"]
        nombre = r["nombre"]
        afp = str(r["afp"]).lower() if r["afp"] else "modelo"
        isapre = str(r["isapre"]).lower() if r["isapre"] else "fonasa"
        tipo_contrato = r["tipo_contrato"]
        sueldo_base = r["sueldo_base"]
        
        # Reconstruct full monthly contracted base salary if the record has proportional days worked
        dias = r["dias_trabajados"] or 30
        if 0 < dias < 30:
            sueldo_base = round(sueldo_base * 30.0 / dias)
            
        isapre_key = "fonasa" if "fona" in isapre.lower() else isapre
        cotiz_uf = 0.0
        cotiz_pesos = 0.0
        if isapre_key != "fonasa":
            cotiz_pesos = r["cotizacion_salud"]
            
        cursor.execute("""
            INSERT OR IGNORE INTO empleados (
                rut, contrato, nombre, afp, cotizacion_afp, isapre, cotizacion_uf, cotizacion_pesos,
                tipo_contrato, sueldo_base, afecto_seguro_cesantia, horas_semanales, tramo_asig_fam, numero_hijos,
                cargo, centro_costo, sede, agrupacion, area, fecha_inicio_contrato, fecha_termino_contrato,
                id_obra, fecha_finiquito
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            rut, contrato, nombre, afp, 0.0, isapre, cotiz_uf, cotiz_pesos,
            tipo_contrato, sueldo_base, 1, 40.0, "D", 0,
            r["cargo"], r["centro_costo"], r["sede"], "", "", r["fecha_inicio"], r["fecha_termino"],
            cc_to_id_obra.get(r["centro_costo"], ""), r["fecha_termino"] if r["fecha_termino"] else None
        ))
        
    # Update existing employees who have null or empty cargo, cc, or sede
    cursor.execute("""
        UPDATE empleados
        SET cargo = (SELECT c.cargo FROM rex_comparisons c WHERE c.rut = empleados.rut AND c.contrato = empleados.contrato AND c.cargo IS NOT NULL AND c.cargo != '' ORDER BY c.periodo DESC LIMIT 1)
        WHERE cargo IS NULL OR cargo = ''
    """)
    cursor.execute("""
        UPDATE empleados
        SET centro_costo = (SELECT c.centro_costo FROM rex_comparisons c WHERE c.rut = empleados.rut AND c.contrato = empleados.contrato AND c.centro_costo IS NOT NULL AND c.centro_costo != '' ORDER BY c.periodo DESC LIMIT 1)
        WHERE centro_costo IS NULL OR centro_costo = ''
    """)
    cursor.execute("""
        UPDATE empleados
        SET sede = (SELECT c.sede FROM rex_comparisons c WHERE c.rut = empleados.rut AND c.contrato = empleados.contrato AND c.sede IS NOT NULL AND c.sede != '' ORDER BY c.periodo DESC LIMIT 1)
        WHERE sede IS NULL OR sede = ''
    """)
    cursor.execute("""
        UPDATE empleados
        SET fecha_inicio_contrato = (SELECT c.fecha_inicio FROM rex_comparisons c WHERE c.rut = empleados.rut AND c.contrato = empleados.contrato AND c.fecha_inicio IS NOT NULL AND c.fecha_inicio != '' ORDER BY c.periodo DESC LIMIT 1)
        WHERE fecha_inicio_contrato IS NULL OR fecha_inicio_contrato = ''
    """)
    cursor.execute("""
        UPDATE empleados
        SET fecha_termino_contrato = (SELECT c.fecha_termino FROM rex_comparisons c WHERE c.rut = empleados.rut AND c.contrato = empleados.contrato AND c.fecha_termino IS NOT NULL AND c.fecha_termino != '' ORDER BY c.periodo DESC LIMIT 1)
        WHERE fecha_termino_contrato IS NULL OR fecha_termino_contrato = ''
    """)

    # Correct existing employees whose sueldo_base matches the proportional sueldo_base from a proportional month
    cursor.execute("""
        SELECT e.rut, e.contrato, e.sueldo_base as emp_base, c.sueldo_base as comp_base, c.dias_trabajados
        FROM empleados e
        JOIN rex_comparisons c ON e.rut = c.rut AND e.contrato = c.contrato
        WHERE c.dias_trabajados > 0 AND c.dias_trabajados < 30 AND e.sueldo_base = c.sueldo_base
    """)
    to_fix = cursor.fetchall()
    for row in to_fix:
        full_base = round(row["comp_base"] * 30.0 / row["dias_trabajados"])
        cursor.execute("UPDATE empleados SET sueldo_base = ? WHERE rut = ? AND contrato = ?", (full_base, row["rut"], row["contrato"]))
        
    # Update sueldo_base in empleados based on the latest rex_comparisons where sueldo_base > 0
    cursor.execute("""
        SELECT r.rut, r.contrato, r.sueldo_base, r.dias_trabajados
        FROM rex_comparisons r
        JOIN (
            SELECT rut, contrato, MAX(periodo) as max_periodo
            FROM rex_comparisons
            WHERE sueldo_base > 0
            GROUP BY rut, contrato
        ) latest ON r.rut = latest.rut AND r.contrato = latest.contrato AND r.periodo = latest.max_periodo
        WHERE r.sueldo_base > 0
    """)
    latest_sueldos = cursor.fetchall()
    updated_sueldo_count = 0
    for rut, contrato, sb, dias in latest_sueldos:
        if dias and dias > 0 and dias < 30:
            sb = round(sb * 30.0 / dias)
        cursor.execute("""
            UPDATE empleados
            SET sueldo_base = ?
            WHERE rut = ? AND contrato = ? AND (sueldo_base != ? OR sueldo_base IS NULL)
        """, (sb, rut, contrato, sb))
        updated_sueldo_count += cursor.rowcount
    print(f"[Migration] Updated sueldo_base for {updated_sueldo_count} employees based on latest comparisons.")

    # Parse and register the Nomina Finiquitos Excel file
    try:
        import glob
        files = glob.glob(os.path.join(BASE_DIR, "Nomina Finiquitos *.xlsx"))
        for filepath in files:
            print(f"Parsing Nomina Finiquitos file: {os.path.basename(filepath)}")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'Detalle' not in wb.sheetnames:
                print(f"Skipping {filepath} - no 'Detalle' sheet.")
                continue
            sheet = wb['Detalle']
            
            # Find headers "Número ID", "CONTRATO"
            first_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            idx_rut = next((i for i, h in enumerate(first_row) if 'número id' in h.lower() or 'numero id' in h.lower()), None)
            idx_contrato = next((i for i, h in enumerate(first_row) if 'contrato' in h.lower()), None)
            
            if idx_rut is None or idx_contrato is None:
                print(f"Error: Could not find 'Número ID' or 'CONTRATO' columns in {filepath}.")
                continue
                
            count_f = 0
            for r_idx in range(2, sheet.max_row + 1):
                rut_val = sheet.cell(row=r_idx, column=idx_rut+1).value
                contrato_val = sheet.cell(row=r_idx, column=idx_contrato+1).value
                if not rut_val:
                    continue
                
                rut_clean = clean_rut(rut_val)
                contrato_num = int(contrato_val) if contrato_val is not None else 1
                
                # Get the fecha_termino from rex_comparisons for June 2026
                cursor.execute("""
                    SELECT fecha_termino FROM rex_comparisons 
                    WHERE rut = ? AND contrato = ? AND periodo = '2026-06'
                """, (rut_clean, contrato_num))
                res = cursor.fetchone()
                fin_date = res[0] if res else None
                
                # If not found, use fecha_termino_contrato from empleados
                if not fin_date:
                    cursor.execute("""
                        SELECT fecha_termino_contrato FROM empleados 
                        WHERE rut = ? AND contrato = ?
                    """, (rut_clean, contrato_num))
                    res2 = cursor.fetchone()
                    fin_date = res2[0] if res2 else "2026-06-30"
                    
                if not fin_date or fin_date == '':
                    fin_date = "2026-06-30"
                    
                cursor.execute("""
                    UPDATE empleados
                    SET fecha_finiquito = ?
                    WHERE rut = ? AND contrato = ?
                """, (fin_date, rut_clean, contrato_num))
                count_f += 1
            print(f"[Migration] Registered {count_f} finiquitos from {os.path.basename(filepath)}.")
    except Exception as ex:
        print(f"Error loading Nomina Finiquitos file: {ex}")

    conn.commit()
    conn.close()
    print("Missing historical employees populated and existing employees' details updated.")

def load_vacations_report_from_excel():
    excel_path = os.path.join(BASE_DIR, "Reporte detallado de vacaciones.xlsx")
    if not os.path.exists(excel_path):
        print("Warning: Reporte detallado de vacaciones.xlsx not found!")
        return
        
    print(f"Loading vacations report from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM vacaciones_reporte")
    
    count = 0
    # Headers in row 1: RUT, Nombre colaborador, Contrato, Cantidad de días, Desde, Hasta, Tipo feriado, Día, Periodo Vacacional
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
            
        rut = clean_rut(row[0])
        contrato = int(row[2]) if row[2] is not None else 1
        dias = float(row[3]) if row[3] is not None else 0.0
        desde = format_date(row[4]) if row[4] is not None else None
        hasta = format_date(row[5]) if row[5] is not None else None
        tipo = str(row[6]).strip() if row[6] is not None else ""
        dia_type = str(row[7]).strip() if row[7] is not None else ""
        periodo = str(row[8]).strip() if row[8] is not None else ""
        
        cursor.execute("""
        INSERT INTO vacaciones_reporte (rut, contrato, dias, desde, hasta, tipo, dia_type, periodo)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (rut, contrato, dias, desde, hasta, tipo, dia_type, periodo))
        count += 1
        
    conn.commit()
    conn.close()
    print(f"Loaded {count} vacation report records into database.")

def sync_vacaciones_ajustes_from_report():
    print("Syncing vacaciones_ajustes from vacations report...")
    import datetime
    import json as json_lib
    
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM vacaciones_ajustes")
    
    cursor.execute("SELECT rut, contrato, nombre, fecha_inicio_contrato, raw_json FROM empleados")
    employees = cursor.fetchall()
    
    today = datetime.date.today()
    count = 0
    
    for emp in employees:
        emp_rut = emp["rut"]
        emp_contrato = emp["contrato"]
        emp_start = emp["fecha_inicio_contrato"]
        emp_raw = emp["raw_json"]
        
        # Get custom vacation start date from raw_json
        vac_start_str = None
        if emp_raw:
            try:
                raw = json_lib.loads(emp_raw)
                vac_start_str = raw.get("Fecha inicio vacaciones")
                if vac_start_str:
                    vac_start_str = vac_start_str[:10]
            except Exception:
                pass
                
        vac_start_str = vac_start_str or emp_start
        
        dias_devengados = 0.0
        if vac_start_str:
            try:
                start_date = datetime.datetime.strptime(vac_start_str[:10], "%Y-%m-%d").date()
                if start_date <= today:
                    diff_days = (today - start_date).days
                    dias_devengados = round((diff_days * 1.25) / 30.0, 2)
            except Exception:
                pass
                
        # Reales adjustment: Sum of 'dias' where desde IS NULL
        cursor.execute("""
            SELECT SUM(dias) FROM vacaciones_reporte
            WHERE rut = ? AND contrato = ? AND desde IS NULL AND tipo = 'Normales'
        """, (emp_rut, emp_contrato))
        excel_reales = cursor.fetchone()[0]
        
        # Tomados: Sum of 'dias' where desde IS NOT NULL and tipo = 'Normales'
        cursor.execute("""
            SELECT SUM(dias) FROM vacaciones_reporte
            WHERE rut = ? AND contrato = ? AND desde IS NOT NULL AND tipo = 'Normales'
        """, (emp_rut, emp_contrato))
        excel_tomados = cursor.fetchone()[0]
        
        dias_reales = float(excel_reales) if excel_reales is not None else dias_devengados
        dias_tomados = float(excel_tomados) if excel_tomados is not None else 0.0
        
        fecha_act = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute("""
            INSERT OR REPLACE INTO vacaciones_ajustes (rut, contrato, dias_reales, dias_tomados, fecha_actualizacion, nota)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (emp_rut, emp_contrato, dias_reales, dias_tomados, fecha_act, "Carga inicial desde Reporte detallado de vacaciones.xlsx"))
        count += 1
        
    conn.commit()
    conn.close()
    print(f"Synced {count} employee vacation adjustments from report.")

def setup():
    init_db()
    seed_parameters()
    load_employees_from_excel()
    load_cargos_mapping()
    load_proyectos_mapping()
    load_rex_comparisons()
    load_planilla_entradas()
    fill_missing_employees()
    load_vacations_report_from_excel()
    sync_vacaciones_ajustes_from_report()
    # Also ensure reconciliaciones table is created on existing DB
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reconciliaciones (
        rut TEXT,
        contrato INTEGER,
        periodo TEXT,
        aprobado INTEGER DEFAULT 0,
        nota TEXT,
        usuario TEXT DEFAULT 'Administrador',
        fecha_aprobacion TEXT DEFAULT (datetime('now', 'localtime')),
        PRIMARY KEY (rut, contrato, periodo)
    )
    """)
    conn.commit()
    conn.close()

    print("Database setup complete.")

def get_audit_alerts(periodo=None):
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    if periodo:
        cursor.execute("""
            SELECT a.*, e.nombre 
            FROM alertas_auditoria a
            LEFT JOIN empleados e ON a.rut = e.rut AND a.contrato = e.contrato
            WHERE a.periodo = ?
            ORDER BY a.rut, a.tipo_alerta
        """, (periodo,))
    else:
        cursor.execute("""
            SELECT a.*, e.nombre 
            FROM alertas_auditoria a
            LEFT JOIN empleados e ON a.rut = e.rut AND a.contrato = e.contrato
            ORDER BY a.periodo DESC, a.rut, a.tipo_alerta
        """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

if __name__ == "__main__":
    setup()

